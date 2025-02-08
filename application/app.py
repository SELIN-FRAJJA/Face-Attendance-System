from flask import Flask, request, jsonify, render_template
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/calculate_wage', methods=['POST'])
def calculate_wage():
    data = request.json
    name = data['name']
    start_date = datetime.strptime(data['startDate'], '%Y-%m-%d')
    end_date = datetime.strptime(data['endDate'], '%Y-%m-%d')
    
    # Initialize variables
    total_wages = 0
    total_days = 0

    # Define daily wage
    daily_wage = 500
    attendance_filename = r'E:\Face-Recognition-Attendance-Projects\application\Attendance.xlsx'
    
    # Load the attendance file
    wb = load_workbook(attendance_filename)
    sheet = wb.active

    # Count attendance days within the specified range
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%d-%m-%Y')  # Format date to match Excel

        # Check if the name and date exist in the attendance file
        found = False
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[0].lower() == name.lower() and row[1] == date_str:
                total_days += 1
                found = True
                break

        current_date += timedelta(days=1)

    # Calculate total wages
    total_wages = total_days * daily_wage
    return jsonify({'success': True, 'total_wages': total_wages, 'total_days': total_days})

@app.route('/get_attendance_data', methods=['POST'])
def get_attendance_data():
    data = request.json
    name = data['name']
    start_date = datetime.strptime(data['startDate'], '%Y-%m-%d')
    end_date = datetime.strptime(data['endDate'], '%Y-%m-%d')

    # Load attendance Excel file
    filename = r'E:\Face-Recognition-Attendance-Projects\application\Attendance.xlsx'
    wb = load_workbook(filename)
    sheet = wb.active

    # Prepare data for the chart
    dates = []
    attendance_counts = []
    
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%d-%m-%Y')
        dates.append(date_str)

        # Count attendance for the current date
        count = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[0].lower() == name.lower() and row[1] == date_str:
                count += 1

        attendance_counts.append(count)
        current_date += timedelta(days=1)

    return jsonify({
        'success': True,
        'dates': dates,
        'attendanceCounts': attendance_counts
    })

if __name__ == '__main__':
    app.run(debug=True)

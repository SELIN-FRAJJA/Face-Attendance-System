import os
from datetime import datetime
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def send_late_notification(recipient_email, name, late_time):
    sender_email = "person754sample@gmail.com"  # Your sender Gmail
    sender_password = "ofzx fole mgjy wdak"       # The app password generated for Python
    subject = "Late Attendance Notification"
    body = f"Dear {name},\n\nYou were marked late at {late_time}. Please ensure timely attendance.\n\nBest regards,\nAttendance System"

    # Create the email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    # Send the email
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()  # Start TLS encryption
        server.login(sender_email, sender_password)  # Login with email and app password
        server.send_message(msg)
        server.quit()
        print(f"Email sent successfully to {recipient_email}!")
    except Exception as e:
        print(f"Failed to send email: {e}")

def check_late_arrivals():
    attendance_filename = r'E:\Face-Recognition-Attendance-Projects\application\Attendance.xlsx'
    email_filename = r'E:\Face-Recognition-Attendance-Projects\application\Attendance_mail.xlsx'
    
    # Load attendance data
    attendance_wb = load_workbook(attendance_filename)
    attendance_sheet = attendance_wb.active

    # Load email data
    email_wb = load_workbook(email_filename)
    email_sheet = email_wb.active

    late_threshold = datetime.strptime("09:00 AM", "%I:%M %p").time()
    
    # Check each entry in the attendance file
    for row in attendance_sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        date = row[1]
        in_out_times = row[2]
        if in_out_times:
            in_time_str = in_out_times.split(' - ')[0]
            in_time = datetime.strptime(in_time_str, '%H:%M:%S').time()

            if in_time > late_threshold:
                # Find the corresponding email address
                for email_row in email_sheet.iter_rows(min_row=2, values_only=True):
                    if email_row[0] == name:  # Match found
                        recipient_email = email_row[1]
                        now = datetime.now()
                        late_time = now.strftime('%I:%M %p')
                        send_late_notification(recipient_email, name, late_time)
                        break

if __name__ == "__main__":
    check_late_arrivals()

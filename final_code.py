#final after adding duration
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

# Path to the training images
path = r'E:\Face-Recognition-Attendance-Projects\Training_images2'
images = []
classNames = []
myList = os.listdir(path)
print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
print(classNames)

def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)
        if len(encode) > 0:
            encodeList.append(encode[0])
    return encodeList

# Dictionary to track the number of times each person has been recognized
recognition_count = {}
# Dictionary to store the last recognition time for each person
recognition_time = {}

def calculate_duration(in_time, out_time):
    in_time = datetime.strptime(in_time, '%H:%M:%S')
    out_time = datetime.strptime(out_time, '%H:%M:%S')
    duration = out_time - in_time
    return str(duration)

def markAttendance(name):
    filename = 'application/Attendance.xlsx'
    
    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Name', 'Date', 'In-Time', 'Out-Time', 'Duration'])
    
    # Get the current date and time
    now = datetime.now()
    dateString = now.strftime('%Y-%m-%d')
    timeString = now.strftime('%H:%M:%S')
    
    if name not in recognition_count:
        recognition_count[name] = 0
    
    recognition_count[name] += 1
    
    # Check the time difference for out-time entry
    if name in recognition_time:
        last_recognition_time = recognition_time[name]
        if (now - last_recognition_time) < timedelta(minutes=1):
            recognition_count[name] -= 1  # Decrement the count if it's within the 1-minute window
            return
    
    # Update the in-time or out-time based on the recognition count
    if recognition_count[name] % 2 == 1:
        sheet.append([name, dateString, timeString, '', ''])  # In-Time
    else:
        # Find the latest entry for the person to update the out-time
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
            if row[0].value == name and row[1].value == dateString and row[3].value is None:
                row[3].value = timeString  # Out-Time
                row[4].value = calculate_duration(row[2].value, timeString)  # Duration
                break
    
    recognition_time[name] = now  # Update the last recognition time
    wb.save(filename)

encodeListKnown = findEncodings(images)
print('Encoding Complete')

cap = cv2.VideoCapture(0)  # Use 0 for the primary webcam

if not cap.isOpened():
    print("Error: Could not open webcam.")
    exit()

while True:
    success, img = cap.read()
    if not success:
        print("Error: Could not read frame from webcam.")
        break

    imgS = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
            markAttendance(name)

    cv2.imshow('Webcam', img)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
